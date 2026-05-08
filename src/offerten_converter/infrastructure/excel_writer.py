"""Excel workbook builder – AMP Sport GmbH corporate design."""

from __future__ import annotations

import io
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── AMP Sport Corporate Colours ───────────────────────────────────────────────
NAVY        = "1B2D6B"   # dark navy  (logo primary)
CYAN        = "29B5E8"   # light blue (logo accent)
NAVY_LIGHT  = "D6DFF0"   # very light navy – alternating rows
CYAN_LIGHT  = "E8F6FC"   # very light cyan – alternating rows
WHITE       = "FFFFFF"
GREY_BORDER = "B0BEC5"
LIGHT_GREY  = "F5F6FA"   # near-white for label cells

THIN_SIDE   = Side(style="thin", color=GREY_BORDER)
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
NO_BORDER   = Border()

# Path to optional logo (placed at <project_root>/assets/logo.png)
# __file__ = src/offerten_converter/infrastructure/excel_writer.py → 4 levels up = project root
_ASSETS = Path(__file__).resolve().parent.parent.parent.parent / "assets"

# EK/Stk and EK Total are never shown in customer-facing output.
# Conditional columns are omitted when the entire column is empty.
_BASE_COLUMNS = [
    ("Pos",           "pos"),
    ("Bezeichnung",   "product_name"),
    ("SKU",           "sku"),
    ("EAN",           "ean"),
    ("Grösse",        "size"),
    ("Farbe",         "color"),
    ("Kategorie",     "category"),
    ("Bestellt",      "qty"),
    ("Max. verfügbar","available_qty"),
    ("VK/Stk",        "vk_unit_target"),
    ("VK Total",      "vk_target"),
    ("Notizen",       "notes"),
    ("Zusatzinfos",   "extra_fields"),
]

_CONDITIONAL = frozenset({"category", "available_qty", "market_price", "notes", "extra_fields"})


def _col_has_data(df: pd.DataFrame, field: str) -> bool:
    """Return True if the column exists and contains at least one meaningful value."""
    if field not in df.columns:
        return False
    col = df[field].dropna()
    if col.empty:
        return False
    cleaned = col.astype(str).str.strip().str.lower()
    return not cleaned[~cleaned.isin({"", "nan", "none", "0", "{}"})].empty


def _active_columns(df: pd.DataFrame) -> list[tuple[str, str]]:
    return [
        (label, field)
        for label, field in _BASE_COLUMNS
        if field not in _CONDITIONAL or _col_has_data(df, field)
    ]


# ── Style helpers ─────────────────────────────────────────────────────────────

def _apply(cell, font=None, fill=None, alignment=None, border=THIN_BORDER):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border is not None:
        cell.border = border


def _navy_cell(ws, row, col, value, size=11, bold=True, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    _apply(c,
        font=Font(name="Calibri", size=size, bold=bold, color=WHITE),
        fill=PatternFill("solid", fgColor=NAVY),
        alignment=Alignment(horizontal=align, vertical="center", wrap_text=True),
    )
    return c


def _cyan_cell(ws, row, col, value, size=10, bold=False, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    _apply(c,
        font=Font(name="Calibri", size=size, bold=bold, color=NAVY),
        fill=PatternFill("solid", fgColor=CYAN_LIGHT),
        alignment=Alignment(horizontal=align, vertical="center"),
    )
    return c


def _label_cell(ws, row, col, value, size=9):
    """Light grey label cell (key in key-value meta pairs)."""
    c = ws.cell(row=row, column=col, value=value)
    _apply(c,
        font=Font(name="Calibri", size=size, bold=True, color=NAVY),
        fill=PatternFill("solid", fgColor=LIGHT_GREY),
        alignment=Alignment(horizontal="right", vertical="center"),
    )
    return c


def _value_cell(ws, row, col, value, size=9):
    """Cyan value cell (value in key-value meta pairs)."""
    c = ws.cell(row=row, column=col, value=value)
    _apply(c,
        font=Font(name="Calibri", size=size, bold=False, color=NAVY),
        fill=PatternFill("solid", fgColor=CYAN_LIGHT),
        alignment=Alignment(horizontal="left", vertical="center"),
    )
    return c


def _data_cell(ws, row, col, value, alt=False):
    c = ws.cell(row=row, column=col, value=value)
    _apply(c,
        font=Font(name="Calibri", size=10, color="1A1A2E"),
        fill=PatternFill("solid", fgColor=NAVY_LIGHT if alt else WHITE),
        alignment=Alignment(vertical="center"),
    )
    return c


def _merge(ws, row, c1, c2, value, font, fill, align="left"):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row=row, column=c1, value=value)
    _apply(c, font=font, fill=fill,
           alignment=Alignment(horizontal=align, vertical="center", wrap_text=True))
    return c


def _cyan_bar(ws, row, n_cols, height=5):
    ws.row_dimensions[row].height = height
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=CYAN)


# ── Main builder ──────────────────────────────────────────────────────────────

def build_excel(
    df: pd.DataFrame,
    supplier_name: str,
    created_by: str,
    target_currency: str,
    valid_days: int = 30,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Offerte"

    today       = date.today()
    valid_until = today + timedelta(days=valid_days)

    OUTPUT_COLUMNS = _active_columns(df)
    N_COLS = len(OUTPUT_COLUMNS)

    # ── Column layout constants ───────────────────────────────────────────────
    # Logo: cols 1-3 | Company info: cols 4..COMPANY_END | Meta: remaining cols
    LOGO_COLS   = 2
    META_COLS   = 4   # 1 label col + 3 value cols on the right
    COMPANY_END = max(N_COLS - META_COLS, LOGO_COLS + 2)
    META_LBL    = COMPANY_END + 1
    META_VAL_S  = META_LBL + 1
    META_VAL_E  = N_COLS

    ROW_HEIGHTS = [50, 18, 18, 18, 4]  # rows 1-5

    # ── White base for entire header area ────────────────────────────────────
    for r, h in enumerate(ROW_HEIGHTS, start=1):
        ws.row_dimensions[r].height = h
        for col in range(1, N_COLS + 1):
            cell = ws.cell(row=r, column=col)
            cell.fill   = PatternFill("solid", fgColor=WHITE)
            cell.border = NO_BORDER

    # ── Logo block: rows 1-4, cols 1-LOGO_COLS ───────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=4, end_column=LOGO_COLS)
    ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor=WHITE)

    logo_path = _ASSETS / "logo.png"
    if logo_path.exists():
        try:
            from openpyxl.drawing.image import Image as XlImage
            img        = XlImage(str(logo_path))
            img.width  = 190
            img.height = 110
            img.anchor = "A1"
            ws.add_image(img)
        except Exception:
            pass

    # ── Company name – row 1 ─────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=LOGO_COLS + 1, end_row=1, end_column=COMPANY_END)
    c = ws.cell(row=1, column=LOGO_COLS + 1, value="AMP Sport GmbH")
    _apply(c,
        font=Font(name="Calibri", size=20, bold=True, color=NAVY),
        fill=PatternFill("solid", fgColor=WHITE),
        alignment=Alignment(horizontal="left", vertical="bottom"),
        border=NO_BORDER,
    )

    # ── Address lines – rows 2-4 ─────────────────────────────────────────────
    addr_lines = ["Poststrasse 6  |  6302 Zug", "Schweiz", "www.ampsport.ch"]
    for row_i, addr in enumerate(addr_lines, start=2):
        ws.merge_cells(start_row=row_i, start_column=LOGO_COLS + 1, end_row=row_i, end_column=COMPANY_END)
        c = ws.cell(row=row_i, column=LOGO_COLS + 1, value=addr)
        _apply(c,
            font=Font(name="Calibri", size=9, color="5A6A8A"),
            fill=PatternFill("solid", fgColor=WHITE),
            alignment=Alignment(horizontal="left", vertical="center"),
            border=NO_BORDER,
        )

    # ── Meta key-value block – rows 1-4, right side ──────────────────────────
    meta_rows = [
        ("Offerte Nr:", ""),
        ("Datum:", today.strftime("%d.%m.%Y")),
        ("Marke:", supplier_name),
        ("Gültig bis:", valid_until.strftime("%d.%m.%Y")),
    ]
    for row_i, (lbl, val) in enumerate(meta_rows, start=1):
        c_lbl = ws.cell(row=row_i, column=META_LBL, value=lbl)
        _apply(c_lbl,
            font=Font(name="Calibri", size=9, bold=True, color=WHITE),
            fill=PatternFill("solid", fgColor=NAVY),
            alignment=Alignment(horizontal="right", vertical="center"),
            border=NO_BORDER,
        )
        ws.merge_cells(start_row=row_i, start_column=META_VAL_S, end_row=row_i, end_column=META_VAL_E)
        c_val = ws.cell(row=row_i, column=META_VAL_S, value=val)
        _apply(c_val,
            font=Font(name="Calibri", size=9, color=NAVY),
            fill=PatternFill("solid", fgColor="EEF2F7"),
            alignment=Alignment(horizontal="left", vertical="center"),
            border=NO_BORDER,
        )

    # ── Row 5 – cyan accent bar ───────────────────────────────────────────────
    _cyan_bar(ws, 5, N_COLS, height=5)

    # ── Row 6 – column headers ────────────────────────────────────────────────
    DATA_START_ROW = 6
    ws.row_dimensions[DATA_START_ROW].height = 26
    for col_idx, (label, _) in enumerate(OUTPUT_COLUMNS, start=1):
        _navy_cell(ws, DATA_START_ROW, col_idx, label, size=10, bold=True)

    # ── Data rows ─────────────────────────────────────────────────────────────
    df_reset = df.reset_index(drop=True)
    data_row = DATA_START_ROW + 1

    # Pre-compute column letters for formula generation
    _field_col = {field: get_column_letter(i + 1) for i, (_, field) in enumerate(OUTPUT_COLUMNS)}
    _qty_col = _field_col.get("qty")
    _vk_unit_col = _field_col.get("vk_unit_target")

    for row_idx, (_, series) in enumerate(df_reset.iterrows()):
        alt = row_idx % 2 == 1
        ws.row_dimensions[data_row].height = 16
        for col_idx, (_, field) in enumerate(OUTPUT_COLUMNS, start=1):
            if field == "pos":
                value = row_idx + 1
            elif field == "vk_target":
                # Formula: blank when qty empty, otherwise VK/Stk × Bestellt
                value = (
                    f'=IF({_qty_col}{data_row}="","",'
                    f'{_vk_unit_col}{data_row}*{_qty_col}{data_row})'
                    if _vk_unit_col and _qty_col else None
                )
            elif field == "qty":
                # Leave qty blank — customer fills this in
                value = None
            elif field == "extra_fields":
                raw = series.get(field)
                value = (
                    " | ".join(f"{k}: {v}" for k, v in raw.items() if v is not None)
                    if isinstance(raw, dict) and raw else None
                )
            else:
                raw = series.get(field)
                if hasattr(raw, "item"):
                    raw = raw.item()
                value = None if (isinstance(raw, float) and pd.isna(raw)) else raw

            cell = _data_cell(ws, data_row, col_idx, value, alt)

            if field in ("vk_unit_target", "vk_target"):
                cell.number_format = f'#,##0.00\\ "{target_currency}"'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif field == "market_price":
                cell.number_format = '#,##0.00\\ "CHF"'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif field == "margin_actual":
                cell.number_format = '0.00"%"'
                cell.alignment = Alignment(horizontal="right", vertical="center")
                try:
                    v = float(value)
                    fg = "27AE60" if v >= 20 else ("E67E22" if v >= 10 else "E74C3C")
                    cell.font = Font(name="Calibri", size=10, bold=True, color=fg)
                except (TypeError, ValueError):
                    pass
            elif field == "qty":
                # Subtle input cell — customer fills this in
                cell.fill = PatternFill("solid", fgColor="EEF2F7")
                cell.font = Font(name="Calibri", size=10, color=NAVY)
                cell.border = Border(
                    left=Side(style="thin", color=NAVY),
                    right=Side(style="thin", color=NAVY),
                    top=Side(style="thin", color=NAVY),
                    bottom=Side(style="thin", color=NAVY),
                )
                cell.number_format = "0"
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif field == "available_qty":
                cell.number_format = "0"
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif field in ("pos", "sku", "ean"):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif field == "product_name":
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

        data_row += 1

    # ── Total row ─────────────────────────────────────────────────────────────
    total_row = data_row
    ws.row_dimensions[total_row].height = 20
    col_map = {field: idx + 1 for idx, (_, field) in enumerate(OUTPUT_COLUMNS)}

    for col_idx in range(1, N_COLS + 1):
        c = ws.cell(row=total_row, column=col_idx)
        c.fill   = PatternFill("solid", fgColor=NAVY)
        c.font   = Font(name="Calibri", size=10, bold=True, color=WHITE)
        c.border = THIN_BORDER
        c.alignment = Alignment(vertical="center")

    ws.cell(row=total_row, column=1).value = "TOTAL"
    ws.cell(row=total_row, column=1).alignment = Alignment(
        horizontal="center", vertical="center"
    )

    sum_start = DATA_START_ROW + 1
    sum_end   = data_row - 1

    if len(df_reset) > 0:
        qty_letter = get_column_letter(col_map["qty"])
        qty_c = ws.cell(row=total_row, column=col_map["qty"])
        qty_c.value         = f"=SUM({qty_letter}{sum_start}:{qty_letter}{sum_end})"
        qty_c.number_format = "0"
        qty_c.alignment     = Alignment(horizontal="center", vertical="center")

        if "available_qty" in col_map:
            avail_letter = get_column_letter(col_map["available_qty"])
            avail_c = ws.cell(row=total_row, column=col_map["available_qty"])
            avail_c.value         = f"=SUM({avail_letter}{sum_start}:{avail_letter}{sum_end})"
            avail_c.number_format = "0"
            avail_c.alignment     = Alignment(horizontal="center", vertical="center")

        vk_letter = get_column_letter(col_map["vk_target"])
        vk_c = ws.cell(row=total_row, column=col_map["vk_target"])
        vk_c.value         = f"=SUM({vk_letter}{sum_start}:{vk_letter}{sum_end})"
        vk_c.number_format = f'#,##0.00\\ "{target_currency}"'
        vk_c.alignment     = Alignment(horizontal="right", vertical="center")

    # ── Freeze, autofilter, column widths ─────────────────────────────────────
    ws.freeze_panes = ws.cell(row=DATA_START_ROW + 1, column=1)
    ws.auto_filter.ref = (
        f"A{DATA_START_ROW}:{get_column_letter(N_COLS)}{data_row - 1}"
    )

    col_widths = {
        "pos": 5, "product_name": 38, "sku": 14, "ean": 16,
        "size": 8, "color": 13, "category": 14, "qty": 9, "available_qty": 14,
        "ek_unit_target": 12, "ek_target": 13, "margin_actual": 10,
        "vk_unit_target": 12, "vk_target": 13,
        "market_price": 14,
        "currency": 9, "notes": 22, "extra_fields": 32,
    }
    for col_idx, (_, field) in enumerate(OUTPUT_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(field, 12)

    # Widen first LOGO_COLS columns so the logo has enough horizontal space
    for col_idx in range(1, LOGO_COLS + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 8

    # ── Footer ────────────────────────────────────────────────────────────────
    footer_row = total_row + 2
    ws.row_dimensions[footer_row].height = 14
    ws.merge_cells(
        start_row=footer_row, start_column=1,
        end_row=footer_row,   end_column=N_COLS,
    )
    c = ws.cell(
        row=footer_row, column=1,
        value="AMP Sport GmbH  ·  Poststrasse 6, 6302 Zug  ·  www.ampsport.ch",
    )
    c.font      = Font(name="Calibri", size=8, italic=True, color="7F8C8D")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = NO_BORDER

    # Thin top border on footer to separate from total
    ws.row_dimensions[total_row + 1].height = 4

    # ── Page setup ────────────────────────────────────────────────────────────
    ws.page_setup.orientation    = "landscape"
    ws.page_setup.fitToPage      = True
    ws.page_setup.fitToWidth     = 1
    ws.page_setup.fitToHeight    = 0
    ws.print_title_rows          = f"1:{DATA_START_ROW}"  # repeat header on each page

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
