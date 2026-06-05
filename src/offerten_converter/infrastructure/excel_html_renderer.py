"""Render an archived spreadsheet to a standalone HTML preview.

Uses openpyxl (already a dependency) to read cells *and* basic styling, so the
archive preview needs no external converter. ``.xlsx``/``.xlsm`` keep bold,
colours, column widths and merged cells; other formats fall back to a plain
value table via pandas.
"""

from __future__ import annotations

import html as html_lib
import logging
from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from offerten_converter.application.ports import SpreadsheetPreviewRenderer

logger = logging.getLogger(__name__)

_MAX_ROWS = 400
_MAX_COLS = 60


class PreviewRenderError(RuntimeError):
    """Raised when the spreadsheet cannot be read for previewing."""


class ExcelHtmlRenderer(SpreadsheetPreviewRenderer):
    """Render spreadsheet bytes to a self-contained HTML document."""

    def to_html(self, data: bytes, *, src_suffix: str) -> str:
        suffix = (src_suffix or "").lower()
        if suffix in (".xlsx", ".xlsm", ""):
            try:
                return self._render_openpyxl(data)
            except Exception:  # noqa: BLE001 - fall back to a plain table
                logger.warning("openpyxl preview failed, falling back to pandas", exc_info=True)
        return self._render_pandas(data)

    # -- openpyxl (styled) --------------------------------------------------

    def _render_openpyxl(self, data: bytes) -> str:
        wb = load_workbook(BytesIO(data), data_only=True)
        ws: Worksheet = wb.active
        if ws is None:
            return _document("<p class='empty'>Keine Daten.</p>")

        max_row = min(ws.max_row or 0, _MAX_ROWS)
        max_col = min(ws.max_column or 0, _MAX_COLS)
        if max_row == 0 or max_col == 0:
            return _document("<p class='empty'>Leeres Tabellenblatt.</p>")

        covered, anchors = _merged_layout(ws, max_row, max_col)

        cols = "".join(
            f"<col style='width:{_col_width_px(ws, c)}px'>" for c in range(1, max_col + 1)
        )
        rows_html: list[str] = []
        for r in range(1, max_row + 1):
            cells_html: list[str] = []
            for c in range(1, max_col + 1):
                if (r, c) in covered:
                    continue
                cell = ws.cell(row=r, column=c)
                span = anchors.get((r, c))
                cells_html.append(_td(cell, span))
            rows_html.append(f"<tr>{''.join(cells_html)}</tr>")

        sheet_count = len(wb.sheetnames)
        note = (
            f"<p class='note'>Blatt „{html_lib.escape(ws.title)}“"
            + (f" (von {sheet_count})" if sheet_count > 1 else "")
            + "</p>"
        )
        truncated = ""
        if (ws.max_row or 0) > _MAX_ROWS or (ws.max_column or 0) > _MAX_COLS:
            truncated = "<p class='note'>Vorschau gekürzt (sehr grosse Tabelle).</p>"

        table = f"<table><colgroup>{cols}</colgroup><tbody>{''.join(rows_html)}</tbody></table>"
        return _document(note + table + truncated)

    # -- pandas fallback (plain) -------------------------------------------

    def _render_pandas(self, data: bytes) -> str:
        import pandas as pd

        try:
            df = pd.read_excel(BytesIO(data), sheet_name=0, header=None, nrows=_MAX_ROWS)
        except Exception as exc:  # noqa: BLE001
            raise PreviewRenderError("Datei konnte nicht gelesen werden") from exc
        if df.empty:
            return _document("<p class='empty'>Keine Daten.</p>")
        df = df.fillna("")
        body = "".join(
            "<tr>" + "".join(f"<td>{html_lib.escape(str(v))}</td>" for v in row) + "</tr>"
            for row in df.itertuples(index=False)
        )
        return _document(f"<table><tbody>{body}</tbody></table>")


# --- helpers --------------------------------------------------------------

def _merged_layout(
    ws: Worksheet, max_row: int, max_col: int
) -> tuple[set[tuple[int, int]], dict[tuple[int, int], tuple[int, int]]]:
    """Return covered cells and anchor->(rowspan, colspan) for merged ranges."""
    covered: set[tuple[int, int]] = set()
    anchors: dict[tuple[int, int], tuple[int, int]] = {}
    for rng in ws.merged_cells.ranges:
        if rng.min_row > max_row or rng.min_col > max_col:
            continue
        rowspan = min(rng.max_row, max_row) - rng.min_row + 1
        colspan = min(rng.max_col, max_col) - rng.min_col + 1
        if rowspan > 1 or colspan > 1:
            anchors[(rng.min_row, rng.min_col)] = (rowspan, colspan)
        for r in range(rng.min_row, min(rng.max_row, max_row) + 1):
            for c in range(rng.min_col, min(rng.max_col, max_col) + 1):
                if (r, c) != (rng.min_row, rng.min_col):
                    covered.add((r, c))
    return covered, anchors


def _col_width_px(ws: Worksheet, col: int) -> int:
    letter = get_column_letter(col)
    dim = ws.column_dimensions.get(letter)
    width = getattr(dim, "width", None) if dim else None
    if not width:
        width = 9
    return int(round(width * 7)) + 8


def _td(cell: Cell, span: tuple[int, int] | None) -> str:
    text = _format_value(cell.value, cell.number_format)
    styles = _cell_styles(cell)
    attrs = ""
    if span:
        rowspan, colspan = span
        if rowspan > 1:
            attrs += f" rowspan='{rowspan}'"
        if colspan > 1:
            attrs += f" colspan='{colspan}'"
    style_attr = f" style='{';'.join(styles)}'" if styles else ""
    return f"<td{attrs}{style_attr}>{text}</td>"


def _cell_styles(cell: Cell) -> list[str]:
    styles: list[str] = []
    font = cell.font
    if font:
        if font.bold:
            styles.append("font-weight:700")
        if font.italic:
            styles.append("font-style:italic")
        rgb = _rgb(getattr(font.color, "rgb", None)) if font.color else None
        if rgb and rgb != "000000":
            styles.append(f"color:#{rgb}")
    fill = cell.fill
    if fill and fill.patternType == "solid":
        rgb = _rgb(getattr(fill.fgColor, "rgb", None))
        if rgb and rgb.upper() not in ("FFFFFF", "000000"):
            styles.append(f"background:#{rgb}")
    align = cell.alignment
    if align and align.horizontal in ("center", "right", "left"):
        styles.append(f"text-align:{align.horizontal}")
    elif isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
        styles.append("text-align:right")
    return styles


def _rgb(value: object) -> str | None:
    """Normalise an openpyxl colour (often 'FFRRGGBB') to 'RRGGBB'."""
    if not isinstance(value, str) or len(value) < 6:
        return None
    return value[-6:]


def _format_value(value: object, number_format: str | None) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Ja" if value else "Nein"
    if isinstance(value, (datetime, date)):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, (int, float)):
        fmt = number_format or ""
        if "%" in fmt:
            return f"{value * 100:g}%"
        if float(value).is_integer():
            return html_lib.escape(f"{int(value)}")
        return html_lib.escape(f"{value:.2f}")
    return html_lib.escape(str(value))


def _document(body: str) -> str:
    return (
        "<!doctype html><html lang='de'><head><meta charset='utf-8'>"
        "<style>"
        "body{margin:0;padding:16px;font-family:'IBM Plex Sans',Arial,sans-serif;"
        "font-size:13px;color:#131720;background:#fff;}"
        "table{border-collapse:collapse;}"
        "td{border:1px solid #d7dce5;padding:4px 8px;white-space:nowrap;"
        "overflow:hidden;text-overflow:ellipsis;vertical-align:middle;}"
        ".note{color:#8a94a6;font-size:12px;margin:0 0 10px;}"
        ".empty{color:#8a94a6;}"
        "</style></head><body>" + body + "</body></html>"
    )
