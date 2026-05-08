"""Read supplier offer files into DataFrames with import-time normalization."""

from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass, field
from typing import Any

import pandas as pd

logger = logging.getLogger(__name__)

_TEXT_SIZES = frozenset([
    "XXS", "XS", "S", "M", "L", "XL", "XXL", "XXXL",
    "2XL", "3XL", "4XL", "5XL", "ONE SIZE",
])

_CANONICAL_ALIASES: dict[str, tuple[str, ...]] = {
    "sku": (
        "sku", "code", "reference", "references", "internal ref", "ident",
        "style", "styles", "style number", "artikelnummer", "artikelnr",
    ),
    "ean": ("ean", "ean code", "upc", "upc/ean", "barcode", "gtin"),
    "product_name": (
        "product name", "name", "item-name", "item name", "style name",
        "model-color", "description", "product/color", "bezeichnung",
    ),
    "size": (
        "size", "item size (eu)", "item size (us)", "item size (uk)",
        "size / size run", "size run", "shoe size", "schuhgrösse", "schuhgroesse",
        "größe", "grösse", "groesse",
        "eu size", "us size", "uk size",
    ),
    "color": (
        "color", "colour", "color name", "color description", "color/code color",
        "farbe",
    ),
    "category": (
        "category", "subcategory", "sub-family", "family", "concept", "use for",
        "collection", "kollektion", "kollection", "season", "saison",
        "line", "product line", "genre", "segment", "series",
    ),
    "available_qty": (
        "available qty", "available quantity", "available q.ty", "stock",
        "qty", "max qty", "quantity available", "bestand", "lager",
        "max. verfügbar", "max verfügbar", "max. verfuegbar", "max verfuegbar",
        "verfügbar", "verfuegbar",
    ),
    "ordered_qty": ("order", "ordered qty", "ordered quantity", "order qty", "bestellung"),
    "unit_price": (
        "offer price", "offer price eur€", "offer price eur", "whs", "whs (eur)",
        "wholesale", "exw bcn", "deal", "net price", "unit price",
        "ek/stk", "ek stk", "ek pro stk", "ek/stück", "ek stück",
        "einkaufspreis", "nettopreis",
    ),
    "currency": ("currency", "währung", "waehrung"),
    "discount_pct": ("retail discount %", "discount", "rabatt"),
}

_PRICE_FALLBACK_ALIASES = (
    "rrp", "rrp (eur)", "retail price", "retail price eur€", "retail price eur"
)


@dataclass
class ReadResult:
    """Normalized import result for one supplier offer file."""

    df: pd.DataFrame
    metadata_hints: dict[str, Any] = field(default_factory=dict)
    was_unpivoted: bool = False
    unpivot_info: str = ""


def _detect_header_row(df_raw: pd.DataFrame, max_scan: int = 50) -> int:
    """Find the best header row using a scoring approach."""
    n_cols = len(df_raw.columns)
    min_fill = max(2, int(n_cols * 0.4))
    best_row = 0
    best_score = -1.0

    for i in range(min(len(df_raw), max_scan)):
        row = df_raw.iloc[i]
        non_null = [
            v for v in row
            if v is not None and not (isinstance(v, float) and pd.isna(v))
        ]
        if len(non_null) < min_fill:
            continue

        str_vals = [v for v in non_null if isinstance(v, str) and str(v).strip()]
        unique_str = {str(v).strip().lower() for v in str_vals}
        header_keywords = {
            "sku", "ean", "name", "artikel", "price", "preis", "qty", "menge",
            "color", "farbe", "size", "grösse", "groesse", "description",
            "bezeichnung", "quantity", "discount", "rabatt", "upc", "gtin",
            "num", "id", "style", "gender", "family", "wholesale", "retail",
        }
        keyword_hits = sum(1 for word in unique_str if any(kw in word for kw in header_keywords))
        str_ratio = len(str_vals) / len(non_null) if non_null else 0
        uniqueness = len(unique_str) / len(str_vals) if str_vals else 0
        score = str_ratio * 0.4 + uniqueness * 0.3 + min(keyword_hits / 3, 1.0) * 0.3

        if score > best_score:
            best_score = score
            best_row = i
        if score >= 0.7:
            return i

    return best_row


def get_sheet_names(file_bytes: bytes, filename: str) -> list[str]:
    """Return sheet names for an Excel file. Empty list for CSV."""
    lower = filename.lower()
    if lower.endswith(".csv"):
        return []
    try:
        if lower.endswith(".xlsx"):
            import openpyxl

            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
            names = wb.sheetnames
            wb.close()
            return names
        xl = pd.ExcelFile(io.BytesIO(file_bytes))
        return xl.sheet_names
    except Exception as exc:
        logger.error("Could not read sheet names: %s", exc)
        return []


def get_recommended_sheet_name(file_bytes: bytes, filename: str) -> str | None:
    """Return the sheet that looks most like a machine-readable offer table."""
    lower = filename.lower()
    if lower.endswith(".csv"):
        return None
    try:
        xl = pd.ExcelFile(io.BytesIO(file_bytes))
        scores: list[tuple[float, str]] = []
        for sheet in xl.sheet_names:
            try:
                df_raw = xl.parse(sheet, header=None, nrows=80)
            except Exception:
                continue
            scores.append((_score_sheet(df_raw, sheet), sheet))
        if not scores:
            return xl.sheet_names[0] if xl.sheet_names else None
        return max(scores, key=lambda item: item[0])[1]
    except Exception as exc:
        logger.warning("Could not recommend sheet: %s", exc)
        names = get_sheet_names(file_bytes, filename)
        return names[0] if names else None


def read_offer_file(
    file_bytes: bytes,
    filename: str,
    sheet_name: str | None = None,
) -> ReadResult:
    """Read and normalize an uploaded supplier offer file."""
    lower = filename.lower()
    if lower.endswith(".csv"):
        return ReadResult(df=_read_csv(file_bytes))

    try:
        processed_bytes = (
            _unmerge_cells(file_bytes, sheet_name)
            if lower.endswith(".xlsx")
            else file_bytes
        )
        xl = pd.ExcelFile(io.BytesIO(processed_bytes))
        chosen = sheet_name or get_recommended_sheet_name(file_bytes, filename) or xl.sheet_names[0]
        df_raw = xl.parse(chosen, header=None)

        block_df = _read_repeated_header_blocks(df_raw)
        if block_df is not None:
            metadata_hints = {
                "layout_type": "repeated_product_blocks",
                "source_sheet": chosen,
                "column_mapping": _column_mapping_hint(block_df),
            }
            format_currency = _detect_currency_from_formats(file_bytes, chosen)
            if format_currency:
                metadata_hints["detected_currency"] = format_currency
                block_df["currency"] = format_currency
            logger.info(
                "Read repeated block offer: %d rows, %d columns",
                len(block_df),
                len(block_df.columns),
            )
            return ReadResult(df=block_df, metadata_hints=metadata_hints)

        header_row = _detect_header_row(df_raw)
        metadata_hints = _extract_metadata_hints(df_raw, header_row)
        metadata_hints["source_sheet"] = chosen

        df = xl.parse(chosen, header=header_row, dtype=str)
        df = df.dropna(how="all").reset_index(drop=True)
        df = df[[
            c for c in df.columns
            if not (str(c).startswith("Unnamed:") and df[c].isna().all())
        ]]

        size_cols, other_cols = _detect_size_columns(df)
        was_unpivoted = False
        unpivot_info = ""
        if size_cols:
            original_rows = len(df)
            df = _unpivot_sizes(df, size_cols, other_cols)
            was_unpivoted = True
            unpivot_info = (
                f"Grössen-Spalten erkannt ({len(size_cols)} Grössen: "
                f"{size_cols[0]}-{size_cols[-1]}). "
                f"{original_rows} Zeilen -> {len(df)} Varianten-Zeilen (Unpivot)."
            )
            metadata_hints["layout_type"] = "size_matrix_columns"
        else:
            metadata_hints["layout_type"] = "flat_variant_rows"

        df, mapping = _add_canonical_columns(df)
        df = _drop_non_product_rows(df)
        if mapping:
            metadata_hints["column_mapping"] = mapping

        explicit_currency = _dominant_currency(df["currency"]) if "currency" in df.columns else None
        format_currency = _detect_currency_from_formats(file_bytes, chosen)
        if explicit_currency:
            metadata_hints["detected_currency"] = explicit_currency
        elif format_currency and not metadata_hints.get("detected_currency"):
            metadata_hints["detected_currency"] = format_currency
        if metadata_hints.get("detected_currency") and "currency" not in df.columns:
            df["currency"] = metadata_hints["detected_currency"]

        logger.info("Read file: %d rows, %d columns", len(df), len(df.columns))
        return ReadResult(
            df=df,
            metadata_hints=metadata_hints,
            was_unpivoted=was_unpivoted,
            unpivot_info=unpivot_info,
        )
    except Exception as exc:
        raise ValueError(f"Datei konnte nicht gelesen werden: {exc}") from exc


def read_excel(
    file_bytes: bytes,
    filename: str,
    sheet_name: str | None = None,
) -> pd.DataFrame:
    """Backward-compatible wrapper returning only the imported DataFrame."""
    return read_offer_file(file_bytes, filename, sheet_name).df


def _read_csv(file_bytes: bytes) -> pd.DataFrame:
    for sep in (",", ";", "\t"):
        try:
            df = pd.read_csv(io.BytesIO(file_bytes), sep=sep, dtype=str)
            if len(df.columns) > 1:
                return df
        except Exception:
            continue
    return pd.read_csv(io.BytesIO(file_bytes), dtype=str)


def _normalize_label(value: object) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def _has_non_empty_values(series: pd.Series) -> bool:
    values = series.dropna().astype(str).str.strip()
    values = values[~values.str.lower().isin(["", "nan", "none"])]
    return not values.empty


def _parse_price_from_text(value: object) -> str | None:
    text = str(value or "")
    if not text.strip():
        return None
    match = re.search(
        r"(?:rrp|retail price|uvp)\s*[:=]?\s*(?:CHF|EUR|USD|€|\$)?\s*"
        r"(\d{1,6}(?:[.,]\d{1,2})?)",
        text,
        re.IGNORECASE,
    )
    if not match:
        return None
    return match.group(1).replace(",", ".")


def _drop_non_product_rows(df: pd.DataFrame) -> pd.DataFrame:
    """Drop obvious totals/footers after canonical product columns are available."""
    if df.empty:
        return df

    keep = pd.Series(True, index=df.index)
    for col in df.columns:
        if _normalize_label(col) in ("pos", "position"):
            pos_text = df[col].fillna("").astype(str).map(_normalize_label)
            keep &= ~pos_text.isin({"total", "subtotal", "summe", "gesamt"})

    identity_cols = [col for col in ("product_name", "sku", "ean") if col in df.columns]
    if identity_cols:
        identity = df[identity_cols].fillna("").astype(str).agg(" ".join, axis=1).map(
            _normalize_label
        )
        keep &= identity.str.strip() != ""
        footer_terms = ("www.", "poststrasse", "amp sport gmbh", "offerten converter")
        keep &= ~identity.apply(lambda text: any(term in text for term in footer_terms))

    return df[keep].reset_index(drop=True)


def _dominant_currency(series: pd.Series) -> str | None:
    values = series.dropna().astype(str).str.strip().str.upper()
    values = values[values.str.fullmatch(r"[A-Z]{3}")]
    if values.empty:
        return None
    return str(values.mode().iloc[0])


def _matches_alias(label: object, aliases: tuple[str, ...]) -> bool:
    normalized = _normalize_label(label)
    return any(alias == normalized or alias in normalized for alias in aliases)


def _column_mapping_hint(df: pd.DataFrame) -> dict[str, str]:
    mapping = {}
    for canonical in _CANONICAL_ALIASES:
        if canonical in df.columns:
            mapping[canonical] = canonical
    return mapping


def _score_sheet(df_raw: pd.DataFrame, sheet_name: str) -> float:
    header_row = _detect_header_row(df_raw)
    header_values = [
        _normalize_label(v)
        for v in df_raw.iloc[header_row].tolist()
        if v is not None and str(v).strip()
    ]
    alias_hits = 0
    for aliases in _CANONICAL_ALIASES.values():
        if any(_matches_alias(value, aliases) for value in header_values):
            alias_hits += 1
    size_hits = sum(1 for value in header_values if _looks_like_size_label(value))
    non_empty_rows = int(df_raw.dropna(how="all").shape[0])
    score = alias_hits * 10 + min(non_empty_rows, 500) / 50
    if size_hits >= 3:
        score += min(size_hits, 30)
    if "stock" in sheet_name.lower():
        score += 20
    if "overview" in sheet_name.lower():
        score += 10
    return score


def _add_canonical_columns(df: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, str]]:
    df = df.copy()
    mapping: dict[str, str] = {}
    if "_size_from_col" in df.columns and "size" not in df.columns:
        df["size"] = df["_size_from_col"]
        mapping["size"] = "_size_from_col"
    if "_qty_from_col" in df.columns and "available_qty" not in df.columns:
        df["available_qty"] = df["_qty_from_col"]
        mapping["available_qty"] = "_qty_from_col"

    for canonical, aliases in _CANONICAL_ALIASES.items():
        if canonical in df.columns:
            mapping.setdefault(canonical, canonical)
            continue
        source = _find_column(df, aliases)
        if source is not None:
            df[canonical] = df[source]
            mapping[canonical] = str(source)

    if "unit_price" not in df.columns:
        fallback = _find_column(df, _PRICE_FALLBACK_ALIASES)
        if fallback is not None:
            df["unit_price"] = df[fallback]
            mapping["unit_price"] = str(fallback)
    extra_source = _find_column(df, ("zusatzinfos", "extra fields", "extra_fields", "notes"))
    if extra_source is not None:
        parsed_prices = df[extra_source].map(_parse_price_from_text)
        if _has_non_empty_values(parsed_prices):
            if "unit_price" in df.columns:
                unit_price = df["unit_price"].copy()
                blank_values = ["", "nan", "None"]
                blank = unit_price.isna() | unit_price.astype(str).str.strip().isin(blank_values)
                unit_price.loc[blank] = parsed_prices.loc[blank]
                df["unit_price"] = unit_price
                if blank.any():
                    existing = mapping.get("unit_price", "unit_price")
                    mapping["unit_price"] = f"{existing} + {extra_source} (rrp)"
            else:
                df["unit_price"] = parsed_prices
                mapping["unit_price"] = f"{extra_source} (rrp)"

    return df, mapping


def _find_column(df: pd.DataFrame, aliases: tuple[str, ...]) -> object | None:
    for col in df.columns:
        if _matches_alias(col, aliases):
            return col
    return None


def _looks_like_size_label(value: object) -> bool:
    text = str(value or "").strip().upper()
    if not text:
        return False
    if text in _TEXT_SIZES:
        return True
    # Pure number in shoe/clothing range: 36, 38.5, 10, 10.5
    if re.fullmatch(r"\d+(?:[.,]\d+)?", text):
        try:
            number = float(text.replace(",", "."))
        except ValueError:
            return False
        return 1 <= number <= 55
    # Number + size system suffix: "38 EU", "9.5 US", "8 UK", "42 DE", "42 FR"
    m = re.fullmatch(r"(\d+(?:[.,]\d+)?)\s*(?:EU|US|UK|FR|DE|IT|BR|JP|CN|CM)", text)
    if m:
        try:
            return 1 <= float(m.group(1).replace(",", ".")) <= 55
        except ValueError:
            pass
    # Youth / child: 5Y, 8C
    return bool(re.fullmatch(r"\d+(?:[.,]\d+)?[CY]", text))


def _coerce_qty(value: object) -> float | None:
    text = str(value or "").strip().replace("'", "")
    if not text:
        return None
    match = re.match(r"^(\d+(?:[.,]\d+)?)\+?$", text)
    if not match:
        return None
    return float(match.group(1).replace(",", "."))


def _read_repeated_header_blocks(df_raw: pd.DataFrame) -> pd.DataFrame | None:
    rows: list[dict[str, object]] = []
    for idx in range(len(df_raw)):
        row = df_raw.iloc[idx]
        normalized_cells = [_normalize_label(v) for v in row.tolist()]
        if not (
            any("internal ref" == cell for cell in normalized_cells)
            and any("size" == cell for cell in normalized_cells)
            and any("available qty" == cell for cell in normalized_cells)
        ):
            continue

        header = {
            col_idx: str(value).strip()
            for col_idx, value in row.items()
            if value is not None and str(value).strip()
        }
        context = _block_context(df_raw, idx)
        data_idx = idx + 1
        while data_idx < len(df_raw):
            data_row = df_raw.iloc[data_idx]
            values = {
                label: data_row.iloc[col_idx]
                for col_idx, label in header.items()
                if col_idx < len(data_row)
            }
            sku = values.get("Internal Ref")
            size = values.get("Size")
            available = values.get("Available QTY")
            ordered = values.get("Ordered QTY")
            if not sku or not size:
                break
            if _coerce_qty(available) is None and _coerce_qty(ordered) is None:
                break

            rows.append({
                **context,
                "sku": sku,
                "size": size,
                "unit_price": values.get("WHS (EUR)") or values.get("WHS"),
                "available_qty": available,
                "ordered_qty": ordered,
                "currency": "EUR",
                "extra_fields": {
                    "rrp": values.get("RRP (EUR)"),
                    "subtotal": values.get("Subtotal"),
                    "layout_source": "repeated_product_blocks",
                },
            })
            data_idx += 1

    if len(rows) < 2:
        return None
    return pd.DataFrame(rows)


def _block_context(df_raw: pd.DataFrame, header_idx: int) -> dict[str, object]:
    context: dict[str, object] = {}
    for idx in range(max(0, header_idx - 10), header_idx):
        row = df_raw.iloc[idx]
        first = row.iloc[0] if len(row) else None
        second = row.iloc[1] if len(row) > 1 else None
        third = row.iloc[2] if len(row) > 2 else None
        first_text = str(first or "").strip()
        second_text = str(second or "").strip().lower()
        third_text = str(third or "").strip()

        if first_text and not second_text and "offer" not in first_text.lower():
            context["product_name"] = first_text
        if second_text == "master item:":
            context["master_item"] = third_text
        elif second_text == "masteritem & color:":
            context["master_item_color"] = third_text
        elif second_text == "product/color:":
            context["color"] = third_text
    return context


def _detect_currency_from_formats(file_bytes: bytes, sheet_name: str | None = None) -> str | None:
    """Scan Excel number formats for currency symbols and return an ISO code."""
    try:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        format_map = [
            (r"CHF", "CHF"),
            (r"\bFR\b", "CHF"),
            (r"€|EUR", "EUR"),
            (r"\$(?![\w])", "USD"),
            (r"USD", "USD"),
            (r"£|GBP", "GBP"),
            (r"NOK", "NOK"),
            (r"SEK", "SEK"),
            (r"DKK", "DKK"),
        ]
        formats_seen: dict[str, int] = {}
        for row in ws.iter_rows():
            for cell in row:
                fmt = cell.number_format or ""
                if fmt in ("General", "@", ""):
                    continue
                for pattern, iso in format_map:
                    if re.search(pattern, fmt, re.IGNORECASE):
                        formats_seen[iso] = formats_seen.get(iso, 0) + 1
                        break
        wb.close()
        if not formats_seen:
            return None
        return max(formats_seen, key=lambda key: formats_seen[key])
    except Exception:
        return None


def _unmerge_cells(file_bytes: bytes, sheet_name: str | None = None) -> bytes:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb[sheet_name] if sheet_name else wb.active
    if not ws.merged_cells.ranges:
        wb.close()
        return file_bytes
    for merge_range in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = merge_range.bounds
        top_left_value = ws.cell(min_row, min_col).value
        ws.unmerge_cells(str(merge_range))
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row, col).value = top_left_value
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf.read()


def _extract_metadata_hints(df_raw: pd.DataFrame, header_row: int) -> dict[str, Any]:
    hints: dict[str, Any] = {}
    for i in range(header_row):
        row_text = " ".join(
            str(v) for v in df_raw.iloc[i] if v is not None and str(v).strip()
        ).upper()
        curr_match = re.search(
            r"(?:CURRENCY|WÄHRUNG|WAEHRUNG)\s*[=:]\s*(USD|EUR|CHF|GBP)",
            row_text,
        )
        if curr_match:
            hints["detected_currency"] = curr_match.group(1)
    return hints


def _detect_size_columns(df: pd.DataFrame) -> tuple[list[str], list[str]]:
    size_cols: list[str] = []
    other_cols: list[str] = []
    for col in df.columns:
        if _looks_like_size_label(col):
            size_cols.append(col)
        else:
            other_cols.append(col)
    if len(size_cols) >= 3:
        return size_cols, other_cols
    return [], list(df.columns)


def _unpivot_sizes(df: pd.DataFrame, size_cols: list[str], other_cols: list[str]) -> pd.DataFrame:
    melted = df.melt(
        id_vars=other_cols,
        value_vars=size_cols,
        var_name="_size_from_col",
        value_name="_qty_from_col",
    )
    melted["_qty_from_col"] = melted["_qty_from_col"].map(_coerce_qty)
    melted = melted[melted["_qty_from_col"].notna() & (melted["_qty_from_col"] > 0)]
    return melted.reset_index(drop=True)
