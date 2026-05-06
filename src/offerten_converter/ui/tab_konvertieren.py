"""Tab 1: Konvertieren – Upload → Extract → Price → Export."""

from __future__ import annotations

import io
import logging
import os
import re
from dataclasses import dataclass, field
from datetime import date
from typing import Any

import pandas as pd
import streamlit as st

from offerten_converter.application.calculate_prices import enrich_dataframe
from offerten_converter.application.export_quotation import export_to_excel
from offerten_converter.application.extract_products import extract_line_items
from offerten_converter.application.manage_profiles import profile_to_hints
from offerten_converter.application.sanitize_data import sanitize_dataframe
from offerten_converter.infrastructure.file_profile_repo import FileProfileRepository
from offerten_converter.ui.state import clear_extraction, get_settings

logger = logging.getLogger(__name__)

_repo = FileProfileRepository()

_EAN_ALIASES = frozenset([
    "ean", "upc/ean", "upc", "gtin", "barcode", "ean-code", "ean_code",
    "ean nr", "ean-nr", "artikel ean",
])

# Pricing per 1M tokens (USD)
_PRICE_INPUT_PER_M  = 15.0
_PRICE_OUTPUT_PER_M = 75.0
_CHARS_PER_TOKEN    = 4

# Size detection
_SHOE_SIZE_RANGE = (20, 55)
_TEXT_SIZES = frozenset([
    "XXS", "XS", "S", "M", "L", "XL", "XXL", "XXXL",
    "2XL", "3XL", "4XL", "5XL", "ONE SIZE",
])


@dataclass
class ReadResult:
    df: pd.DataFrame
    metadata_hints: dict[str, Any] = field(default_factory=dict)
    was_unpivoted: bool = False
    unpivot_info: str = ""


# ── File reading helpers ───────────────────────────────────────────────────────

def _detect_header_row(df_raw: pd.DataFrame, max_scan: int = 50) -> int:
    n_cols = len(df_raw.columns)
    min_fill = max(3, int(n_cols * 0.75))
    for i in range(min(len(df_raw), max_scan)):
        row = df_raw.iloc[i]
        non_null = [v for v in row if v is not None and not (isinstance(v, float) and pd.isna(v))]
        if len(non_null) < min_fill:
            continue
        str_vals = [v for v in non_null if isinstance(v, str) and str(v).strip()]
        if len(str_vals) / len(non_null) >= 0.6:
            return i
    return 0


def _unmerge_cells(file_bytes: bytes, sheet_name: str | None = None) -> bytes:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb[sheet_name] if sheet_name else wb.active
    if not ws.merged_cells.ranges:
        return file_bytes
    for merge_range in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = merge_range.bounds
        top_left_value = ws.cell(min_row, min_col).value
        ws.unmerge_cells(str(merge_range))
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                ws.cell(r, c).value = top_left_value
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _extract_metadata_hints(df_raw: pd.DataFrame, header_row: int) -> dict[str, Any]:
    hints: dict[str, Any] = {}
    for i in range(header_row):
        row_text = " ".join(
            str(v) for v in df_raw.iloc[i] if v is not None and str(v).strip()
        ).upper()
        curr_match = re.search(
            r"(?:CURRENCY|WÄHRUNG|WAEHRUNG)\s*[=:]\s*(USD|EUR|CHF|GBP)", row_text,
        )
        if curr_match:
            hints["detected_currency"] = curr_match.group(1)
    return hints


def _detect_size_columns(df: pd.DataFrame) -> tuple[list[str], list[str]]:
    size_cols: list[str] = []
    other_cols: list[str] = []
    for col in df.columns:
        try:
            val = float(col)
            if _SHOE_SIZE_RANGE[0] <= val <= _SHOE_SIZE_RANGE[1]:
                size_cols.append(col)
            else:
                other_cols.append(col)
        except (ValueError, TypeError):
            other_cols.append(col)
    if len(size_cols) >= 3:
        return size_cols, other_cols
    size_cols, other_cols = [], []
    for col in df.columns:
        if str(col).upper().strip() in _TEXT_SIZES:
            size_cols.append(col)
        else:
            other_cols.append(col)
    if len(size_cols) >= 3:
        return size_cols, other_cols
    return [], list(df.columns)


def _unpivot_sizes(df: pd.DataFrame, size_cols: list[str], other_cols: list[str]) -> pd.DataFrame:
    melted = df.melt(id_vars=other_cols, value_vars=size_cols,
                     var_name="_size_from_col", value_name="_qty_from_col")
    melted["_qty_from_col"] = pd.to_numeric(melted["_qty_from_col"], errors="coerce")
    melted = melted[melted["_qty_from_col"].notna() & (melted["_qty_from_col"] > 0)]
    return melted.reset_index(drop=True)


def _read_uploaded_file(uploaded_file) -> ReadResult:
    name = uploaded_file.name.lower()
    raw_bytes = uploaded_file.read()

    if name.endswith(".csv"):
        for sep in (",", ";", "\t"):
            try:
                df = pd.read_csv(io.BytesIO(raw_bytes), sep=sep, dtype=str)
                if len(df.columns) > 1:
                    return ReadResult(df=df)
            except Exception:
                continue
        return ReadResult(df=pd.read_csv(io.BytesIO(raw_bytes), dtype=str))

    try:
        if name.endswith(".xlsx"):
            import openpyxl
            wb_check = openpyxl.load_workbook(io.BytesIO(raw_bytes))
            ws_check = wb_check.active
            has_merged = bool(ws_check.merged_cells.ranges)
            sheet_names = wb_check.sheetnames
            wb_check.close()
        else:
            has_merged = False
            xl_tmp = pd.ExcelFile(io.BytesIO(raw_bytes))
            sheet_names = xl_tmp.sheet_names

        chosen = sheet_names[0]
        if len(sheet_names) > 1:
            chosen = st.selectbox("Tabellenblatt auswählen:", sheet_names, key="sheet_select")

        if has_merged:
            processed_bytes = _unmerge_cells(raw_bytes, chosen)
            st.caption("📐 Zusammengeführte Zellen wurden aufgelöst.")
        else:
            processed_bytes = raw_bytes

        xl = pd.ExcelFile(io.BytesIO(processed_bytes))
        df_raw = xl.parse(chosen, header=None)
        header_row = _detect_header_row(df_raw)
        metadata_hints = _extract_metadata_hints(df_raw, header_row)

        df = xl.parse(chosen, header=header_row, dtype=str)
        df = df.dropna(how="all").reset_index(drop=True)
        df = df[[c for c in df.columns
                  if not (str(c).startswith("Unnamed:") and df[c].isna().all())]]

        size_cols, other_cols = _detect_size_columns(df)
        was_unpivoted = False
        unpivot_info = ""
        if size_cols:
            original_rows = len(df)
            df = _unpivot_sizes(df, size_cols, other_cols)
            was_unpivoted = True
            unpivot_info = (
                f"{len(size_cols)} Grössen-Spalten erkannt "
                f"({size_cols[0]}–{size_cols[-1]}): "
                f"{original_rows} Zeilen → {len(df)} Varianten."
            )

        return ReadResult(df=df, metadata_hints=metadata_hints,
                          was_unpivoted=was_unpivoted, unpivot_info=unpivot_info)

    except Exception as exc:
        st.error(f"Fehler beim Lesen der Datei: {exc}")
        st.stop()


# ── Market price helpers ───────────────────────────────────────────────────────

def _fetch_market_prices_once(extracted_df: pd.DataFrame) -> None:
    if st.session_state.get("market_prices_fetched"):
        return
    st.session_state["market_prices_fetched"] = True

    ean_col = None
    if "ean" in extracted_df.columns:
        ean_col = "ean"
    else:
        for col in extracted_df.columns:
            if col.strip().lower() in _EAN_ALIASES:
                ean_col = col
                break
    if ean_col is None:
        return

    eans = (extracted_df[ean_col].dropna().astype(str).str.strip()
            .replace({"nan": "", "None": ""}))
    unique_eans = [e for e in eans.unique() if e]
    if not unique_eans:
        return

    try:
        from offerten_converter.infrastructure.market_price_scraper import ToppreiseScraper
    except ImportError:
        return

    scraper = ToppreiseScraper()
    prices: dict[str, float] = {}
    progress = st.progress(0, text="Marktpreise werden geladen …")
    status = st.empty()

    import time
    for i, ean in enumerate(unique_eans):
        status.caption(f"EAN {ean} ({i + 1}/{len(unique_eans)})")
        price = scraper.fetch_price(ean)
        if price is not None:
            prices[ean] = price
        progress.progress((i + 1) / len(unique_eans))
        if i < len(unique_eans) - 1:
            time.sleep(2.5)

    progress.empty()
    status.empty()
    st.session_state["market_prices"] = prices
    if prices:
        st.success(f"✅ Marktpreise geladen: {len(prices)}/{len(unique_eans)} EANs gefunden.")


# ── API cost estimate ──────────────────────────────────────────────────────────

def _api_cost_caption(text: str, n_chunks: int) -> None:
    from offerten_converter.application.extract_products import SYSTEM_PROMPT, _calculate_chunk_size
    system_tokens   = len(SYSTEM_PROMPT) // _CHARS_PER_TOKEN
    content_tokens  = len(text) // _CHARS_PER_TOKEN
    chunk_tokens    = content_tokens // max(n_chunks, 1)
    total_input     = (system_tokens + chunk_tokens) * n_chunks
    lines           = text.splitlines()
    data_lines      = [l for l in lines[1:] if l.strip()]
    sample          = data_lines[:20]
    compressed      = [re.sub(r" {2,}", " ", l).strip() for l in sample]
    avg_chars       = sum(len(l) for l in compressed) / max(len(compressed), 1)
    total_output    = int(len(data_lines) * avg_chars * 3.0 / _CHARS_PER_TOKEN)
    cost_usd        = (total_input / 1_000_000 * _PRICE_INPUT_PER_M
                       + total_output / 1_000_000 * _PRICE_OUTPUT_PER_M)
    cost_chf        = cost_usd * 0.89
    st.caption(f"Geschätzte API-Kosten: ca. **CHF {cost_chf:.3f}**")


# ── Main render ────────────────────────────────────────────────────────────────

def render():
    settings = get_settings()

    # ── STEP 1: Lieferant & Datei ─────────────────────────────────────────────
    st.subheader("1.  Lieferant & Datei")

    col_sup, col_prof = st.columns(2)
    with col_sup:
        supplier_name = st.text_input(
            "Lieferant",
            value=st.session_state["supplier_name"],
            placeholder="z.B. Sport Muster GmbH",
        )
        st.session_state["supplier_name"] = supplier_name

    with col_prof:
        profiles = _repo.list_profiles()
        if profiles:
            load_choice = st.selectbox(
                "Profil laden (optional)",
                ["– kein Profil –"] + profiles,
                key="profile_load_select",
            )
        else:
            load_choice = "– kein Profil –"

    loaded_profile = None
    if load_choice != "– kein Profil –":
        loaded_profile = _repo.load(load_choice)
        if loaded_profile and not supplier_name:
            supplier_name = loaded_profile["name"]
            st.session_state["supplier_name"] = supplier_name
        if loaded_profile:
            st.caption(
                f"Profil: {loaded_profile['name']} · "
                f"Währung: {loaded_profile.get('typical_currency', '?')} · "
                f"Rabatt: {loaded_profile.get('typical_discount', 0)}%"
            )

    uploaded = st.file_uploader(
        "Datei hochladen (.xlsx, .xls, .csv)",
        type=["xlsx", "xls", "csv"],
        key="file_uploader",
    )

    if uploaded is None:
        clear_extraction()
        return

    fingerprint = f"{uploaded.name}::{uploaded.size}"
    if st.session_state["_file_fingerprint"] != fingerprint:
        clear_extraction()
        st.session_state["_file_fingerprint"] = fingerprint

    result      = _read_uploaded_file(uploaded)
    df_raw      = result.df

    file_info = f"{uploaded.name} · {len(df_raw)} Zeilen · {len(df_raw.columns)} Spalten"
    if result.was_unpivoted:
        file_info += f" · {result.unpivot_info}"
    if result.metadata_hints.get("detected_currency"):
        file_info += f" · Währung erkannt: {result.metadata_hints['detected_currency']}"
    st.caption(f"✅ {file_info}")

    st.divider()

    # ── STEP 2: Produkte erkennen ─────────────────────────────────────────────
    st.subheader("2.  Produkte erkennen")

    # Sanitize runs silently
    df_clean, sanitize_log = sanitize_dataframe(df_raw)
    st.session_state["sanitize_log"] = sanitize_log
    removed = [e for e in sanitize_log if e.startswith("Spalte")]
    if removed:
        st.caption(f"🛡️ {len(removed)} sensible Spalte(n) vor Übermittlung entfernt.")

    sanitized_text = df_clean.to_string(index=False)

    from offerten_converter.application.extract_products import _split_table_into_chunks
    n_chunks = len(_split_table_into_chunks(sanitized_text))

    col_btn, col_cost = st.columns([2, 3])
    with col_btn:
        extract_btn = st.button(
            "🔍  Produkte erkennen", type="primary", use_container_width=True,
        )
    with col_cost:
        st.write("")
        _api_cost_caption(sanitized_text, n_chunks)

    if extract_btn:
        for key in ("extracted_df", "raw_api_response", "extraction_error", "api_usage"):
            st.session_state[key] = None

        api_key = os.getenv("ANTHROPIC_API_KEY")
        if not api_key:
            st.error("ANTHROPIC_API_KEY nicht gesetzt. Bitte in der .env-Datei eintragen.")
            return

        hints = profile_to_hints(loaded_profile) if loaded_profile else ""
        extra = []
        if result.was_unpivoted:
            extra.append(
                "Data was unpivoted: '_size_from_col' = size, "
                "'_qty_from_col' = ordered_qty."
            )
        if result.metadata_hints.get("detected_currency"):
            extra.append(f"Detected currency: {result.metadata_hints['detected_currency']}")
        if extra:
            hints = (hints + " | " if hints else "") + " | ".join(extra)

        with st.spinner("Produkte werden erkannt …"):
            try:
                items, usage = extract_line_items(sanitized_text, hints, api_key)
                df_extracted = pd.DataFrame(items)

                if "ordered_qty" in df_extracted.columns:
                    qty_col = pd.to_numeric(df_extracted["ordered_qty"], errors="coerce")
                    if (qty_col > 0).any() and (qty_col == 0).any():
                        n_before = len(df_extracted)
                        df_extracted = df_extracted[qty_col > 0].reset_index(drop=True)
                        st.info(
                            f"{n_before - len(df_extracted)} Positionen mit Menge 0 entfernt."
                        )

                st.session_state["extracted_df"] = df_extracted
                st.session_state["api_usage"]    = usage
                st.success(f"✅ {len(df_extracted)} Produkte erkannt.")
            except ValueError as exc:
                st.session_state["extraction_error"] = str(exc)
            except RuntimeError as exc:
                st.error(f"Fehler: {exc}")
                return

    if st.session_state.get("extraction_error"):
        st.error("Extraktion fehlgeschlagen – bitte erneut versuchen.")
        return

    if st.session_state["extracted_df"] is None:
        return

    # Market prices auto-load (after extraction, once per session)
    _fetch_market_prices_once(st.session_state["extracted_df"])

    st.divider()

    # ── STEP 3: Produkte & Preise ─────────────────────────────────────────────
    st.subheader("3.  Produkte & Preise")

    extracted_df  = st.session_state["extracted_df"]
    rates         = settings["rates"]
    market_prices = st.session_state.get("market_prices", {})

    # Controls: global margin + currency + marktpreis quick-fill
    ctrl1, ctrl2, ctrl3, ctrl4 = st.columns([1.2, 1, 1.5, 1.5])
    margin_pct = ctrl1.number_input(
        "Standard-Marge %",
        min_value=0.0, max_value=99.0,
        value=settings["default_margin"],
        step=0.5, format="%.1f",
        key="margin_input",
    )
    target_currency = ctrl2.selectbox(
        "Währung",
        ["CHF", "EUR", "USD"],
        index=["CHF", "EUR", "USD"].index(settings["default_currency"]),
        key="currency_input",
    )
    apply_mp = False
    mp_pct   = 0.0
    if market_prices:
        mp_pct = ctrl3.number_input(
            "Marktpreis – %",
            min_value=0.0, max_value=99.0,
            value=5.0, step=0.5, format="%.1f",
            key="mp_pct",
            help="VK aller Produkte auf Marktpreis minus X% setzen",
        )
        ctrl4.write("")
        ctrl4.write("")
        apply_mp = ctrl4.button(
            "Auf alle anwenden", key="apply_mp", use_container_width=True,
        )

    # Build enriched base (for EK in target currency)
    enriched_base = enrich_dataframe(extracted_df, margin_pct, target_currency, rates)

    # Build unified pricing table
    rows = []
    for _, row in enriched_base.iterrows():
        ean = str(row.get("ean", "") or "").strip()
        mp  = market_prices.get(ean) if ean else None
        rows.append({
            "Bezeichnung":       row.get("product_name", ""),
            "Grösse":            row.get("size", ""),
            "Farbe":             row.get("color", ""),
            "Menge":             row.get("qty"),
            "EK/Stk":            row.get("ek_unit_target"),
            "Marktpreis 🌐":     mp,
            "Marge %":           margin_pct,
            "VK/Stk (manuell)":  None,
            "Notizen":           str(row.get("notes") or ""),
        })
    unified_df = pd.DataFrame(rows)

    # Restore previous edits
    prev = st.session_state.get("_pricing_edits")
    if prev is not None and len(prev) == len(unified_df):
        for col in ("Menge", "Marge %", "VK/Stk (manuell)", "Notizen"):
            if col in prev.columns:
                unified_df[col] = prev[col].values

    # Apply marktpreis quick-fill
    if apply_mp:
        for i in range(len(unified_df)):
            mp_val = unified_df.at[i, "Marktpreis 🌐"]
            if mp_val is not None and not (isinstance(mp_val, float) and pd.isna(mp_val)):
                unified_df.at[i, "VK/Stk (manuell)"] = round(
                    float(mp_val) * (1 - mp_pct / 100), 2
                )

    has_market = market_prices and any(r["Marktpreis 🌐"] is not None for r in rows)

    col_cfg: dict = {
        "Bezeichnung":      st.column_config.TextColumn(disabled=True, width="large"),
        "Grösse":           st.column_config.TextColumn(disabled=True, width="small"),
        "Farbe":            st.column_config.TextColumn(disabled=True, width="small"),
        "Menge":            st.column_config.NumberColumn(format="%d", width="small"),
        "EK/Stk":           st.column_config.NumberColumn(
                                disabled=True, format="%.2f",
                                label=f"EK/Stk ({target_currency})",
                            ),
        "Marktpreis 🌐":    st.column_config.NumberColumn(
                                disabled=True, format="%.2f",
                                help="Günstigster Preis toppreise.ch – nicht im Export",
                            ) if has_market else None,
        "Marge %":          st.column_config.NumberColumn(
                                min_value=0.0, max_value=99.0, format="%.1f",
                                help="Pro Zeile editierbar",
                            ),
        "VK/Stk (manuell)": st.column_config.NumberColumn(
                                format="%.2f",
                                label=f"VK/Stk ({target_currency})",
                                help="Direkt eingeben – überschreibt Marge %. Leer = automatisch.",
                            ),
        "Notizen":          st.column_config.TextColumn(width="medium"),
    }
    # Remove None entries (hidden columns)
    col_cfg = {k: v for k, v in col_cfg.items() if v is not None}

    display_cols = [c for c in unified_df.columns if c in col_cfg or c == "Marktpreis 🌐"]
    if not has_market and "Marktpreis 🌐" in unified_df.columns:
        display_cols = [c for c in display_cols if c != "Marktpreis 🌐"]

    edited = st.data_editor(
        unified_df[display_cols],
        column_config=col_cfg,
        use_container_width=True,
        num_rows="fixed",
        key="unified_editor",
    )
    st.session_state["_pricing_edits"] = edited.copy()

    # Re-enrich with per-row overrides
    df_for_export = extracted_df.copy()
    for i, erow in edited.iterrows():
        qty_val = erow.get("Menge")
        if qty_val is not None and not (isinstance(qty_val, float) and pd.isna(qty_val)):
            df_for_export.at[i, "ordered_qty"] = qty_val
        note_val = erow.get("Notizen", "")
        if note_val:
            df_for_export.at[i, "notes"] = note_val

    enriched_rows = []
    for i, row in df_for_export.iterrows():
        erow       = edited.iloc[i]
        vk_manual  = erow.get("VK/Stk (manuell)")
        marge_row  = float(erow.get("Marge %") or margin_pct)
        row_df     = pd.DataFrame([row])

        try:
            vk_f = float(vk_manual) if vk_manual is not None else None
            if isinstance(vk_manual, float) and pd.isna(vk_manual):
                vk_f = None
        except (ValueError, TypeError):
            vk_f = None

        if vk_f and vk_f > 0:
            row_df["vk_target"] = vk_f
        enriched_row = enrich_dataframe(row_df, marge_row, target_currency, rates)
        enriched_rows.append(enriched_row)

    enriched = pd.concat(enriched_rows, ignore_index=True)
    st.session_state["enriched_df"] = enriched

    # Summary metrics
    ek_total = enriched["ek_target"].sum(skipna=True)
    vk_total = enriched["vk_target"].sum(skipna=True)
    avg_marg = enriched["margin_actual"].mean(skipna=True)
    m1, m2, m3 = st.columns(3)
    m1.metric(f"EK Total ({target_currency})", f"{ek_total:,.2f}")
    m2.metric(f"VK Total ({target_currency})", f"{vk_total:,.2f}")
    m3.metric("Ø Marge", f"{avg_marg:.1f}%" if not pd.isna(avg_marg) else "–")

    # Advanced options (collapsed)
    with st.expander("Erweiterte Optionen"):
        if supplier_name:
            st.markdown("**Lieferantenprofil speichern**")
            pc1, pc2 = st.columns(2)
            prof_currency = pc1.text_input("Typische Währung", value="EUR", key="prof_cur")
            prof_discount = pc2.number_input(
                "Typischer Rabatt %", 0.0, 100.0, 0.0, 0.5, key="prof_disc",
            )
            if st.button("Profil speichern", key="save_profile"):
                _repo.save(supplier_name, prof_currency, prof_discount, "")
                st.success(f"Profil '{supplier_name}' gespeichert.")

    st.divider()

    # ── STEP 4: Offerte exportieren ───────────────────────────────────────────
    st.subheader("4.  Offerte exportieren")

    effective_supplier = supplier_name or (
        loaded_profile.get("name", "") if loaded_profile else ""
    )
    created_by = settings.get("company_name", "AMP Sport GmbH")
    valid_days = int(settings.get("valid_days", 30))

    if not effective_supplier:
        st.warning("Bitte zuerst den Lieferantennamen eingeben (Schritt 1).")
    elif enriched.empty:
        st.warning("Keine Positionen vorhanden.")
    else:
        try:
            excel_bytes = export_to_excel(
                enriched, effective_supplier, created_by, target_currency, valid_days,
            )
            today_str = date.today().strftime("%Y%m%d")
            filename  = f"Offerte_{effective_supplier.replace(' ', '_')}_{today_str}.xlsx"
            st.download_button(
                label=f"⬇️  Offerte herunterladen  ({len(enriched)} Positionen)",
                data=excel_bytes,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )
        except Exception as exc:
            st.error(f"Export-Fehler: {exc}")
