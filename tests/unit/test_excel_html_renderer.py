"""Unit tests for the openpyxl-based HTML preview renderer."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from offerten_converter.infrastructure.excel_html_renderer import ExcelHtmlRenderer


def _xlsx(build) -> bytes:
    wb = Workbook()
    build(wb.active)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_renders_values_and_bold():
    def build(ws):
        ws["A1"] = "Artikel"
        ws["A1"].font = Font(bold=True)
        ws["B1"] = "Preis"
        ws["A2"] = "Schuh"
        ws["B2"] = 79.9

    html = ExcelHtmlRenderer().to_html(_xlsx(build), src_suffix=".xlsx")
    assert "<table" in html
    assert "Artikel" in html and "Schuh" in html
    assert "79.90" in html  # float formatted to 2 decimals
    assert "font-weight:700" in html  # bold carried over


def test_escapes_cell_content():
    def build(ws):
        ws["A1"] = "<script>alert(1)</script>"

    html = ExcelHtmlRenderer().to_html(_xlsx(build), src_suffix=".xlsx")
    assert "<script>alert(1)</script>" not in html
    assert "&lt;script&gt;" in html


def test_merged_cells_get_colspan():
    def build(ws):
        ws["A1"] = "Titel"
        ws.merge_cells("A1:C1")
        ws["A2"] = "x"

    html = ExcelHtmlRenderer().to_html(_xlsx(build), src_suffix=".xlsx")
    assert "colspan='3'" in html


def test_empty_sheet_is_handled():
    # A blank workbook still has a 1x1 grid; rendering must not raise.
    html = ExcelHtmlRenderer().to_html(_xlsx(lambda ws: None), src_suffix=".xlsx")
    assert "<body>" in html and "</html>" in html
