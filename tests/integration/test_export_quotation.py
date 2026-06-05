"""Integration tests for Excel export."""

import io

import pandas as pd
import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from offerten_converter.infrastructure.excel_writer import _active_columns, build_excel


@pytest.fixture
def sample_df():
    return pd.DataFrame([
        {
            "product_name": "BLACKROLL STANDARD",
            "sku": "BR-001",
            "ean": "4260346270001",
            "size": None,
            "color": "schwarz",
            "category": "Faszienrolle",
            "qty": 10,
            "ek_unit_target": 24.77,
            "currency": "EUR",
            "ek_target": 247.70,
            "margin_actual": 40.0,
            "vk_unit_target": 41.28,
            "vk_target": 412.83,
            "notes": None,
        },
    ])


class TestBuildExcel:
    def test_returns_valid_xlsx(self, sample_df):
        data = build_excel(sample_df, "Test GmbH", "Tester", "CHF", 30)
        assert isinstance(data, bytes)
        assert len(data) > 0
        wb = load_workbook(io.BytesIO(data))
        assert "Offerte" in wb.sheetnames

    def test_has_correct_column_count(self, sample_df):
        data = build_excel(sample_df, "Test GmbH", "Tester", "CHF")
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Offerte"]
        header_row = 6
        filled_cols = sum(
            1 for col in range(1, 50) if ws.cell(row=header_row, column=col).value
        )
        assert filled_cols == len(_active_columns(sample_df))
        assert ws.cell(row=header_row, column=1).value == "Pos"

    def test_meta_header(self, sample_df):
        data = build_excel(sample_df, "Sport AG", "Max", "EUR", 60)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Offerte"]
        # Supplier name is rendered as the "Marke:" value in the meta block
        # (rows 1-4). The meta block sits on the right, so its column depends on
        # the number of active columns – locate it instead of hardcoding.
        # (created_by is currently not rendered in this layout.)
        marke_row = next(
            r for r in range(1, 5)
            if any(ws.cell(row=r, column=c).value == "Marke:" for c in range(1, 50))
        )
        label_col = next(
            c for c in range(1, 50) if ws.cell(row=marke_row, column=c).value == "Marke:"
        )
        assert ws.cell(row=marke_row, column=label_col + 1).value == "Sport AG"

    def test_total_row_exists(self, sample_df):
        data = build_excel(sample_df, "Test", "T", "CHF")
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Offerte"]
        # Header row 6 + 1 data row + TOTAL row => row 8
        total_row = next(
            r for r in range(1, 30) if ws.cell(row=r, column=1).value == "TOTAL"
        )
        assert total_row == 8

    def test_sell_price_columns_labelled_ek(self, sample_df):
        # The offer is the customer's order, so the sell price is labelled "EK".
        data = build_excel(sample_df, "Test", "T", "CHF")
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Offerte"]
        headers = {ws.cell(row=6, column=c).value for c in range(1, 50)}
        assert "EK/Stk" in headers
        assert "EK Total" in headers
        assert "VK/Stk" not in headers
        assert "VK Total" not in headers

    def test_qty_validation_caps_at_available(self):
        df = pd.DataFrame([
            {"product_name": "A", "qty": None, "available_qty": 5,
             "vk_unit_target": 10.0, "vk_target": None},
            {"product_name": "B", "qty": None, "available_qty": 12,
             "vk_unit_target": 20.0, "vk_target": None},
        ])
        data = build_excel(df, "Test", "T", "CHF")
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Offerte"]
        dvs = ws.data_validations.dataValidation
        assert len(dvs) == 1
        dv = dvs[0]
        assert dv.type == "whole"
        assert dv.operator == "between"
        # Relative max reference anchored on the first guarded row.
        assert dv.formula1 == "0"
        avail_col = next(
            get_column_letter(c)
            for c in range(1, 50)
            if ws.cell(row=6, column=c).value == "Max. verfügbar"
        )
        assert dv.formula2 == f"{avail_col}7"
        qty_col = next(
            get_column_letter(c)
            for c in range(1, 50)
            if ws.cell(row=6, column=c).value == "Bestellt"
        )
        assert f"{qty_col}7" in str(dv.sqref)
        assert f"{qty_col}8" in str(dv.sqref)
