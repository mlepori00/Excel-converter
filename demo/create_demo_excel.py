"""Generate a realistic demo supplier quotation Excel file for presentations."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUTPUT_PATH = Path(__file__).parent / "Lieferanten_Offerte_SportPro_Demo.xlsx"

PRODUCTS = [
    # SKU, EAN, Bezeichnung, Kategorie, Farbe, Grösse, Preis EUR, Min.Menge, Rabatt%
    ("SP-RUN-001", "4012345678901", "Running Shoe AirBoost Pro", "Laufschuhe", "Schwarz/Rot", "40", 89.50, 6, 15),
    ("SP-RUN-001", "4012345678902", "Running Shoe AirBoost Pro", "Laufschuhe", "Schwarz/Rot", "41", 89.50, 6, 15),
    ("SP-RUN-001", "4012345678903", "Running Shoe AirBoost Pro", "Laufschuhe", "Schwarz/Rot", "42", 89.50, 6, 15),
    ("SP-RUN-001", "4012345678904", "Running Shoe AirBoost Pro", "Laufschuhe", "Schwarz/Rot", "43", 89.50, 6, 15),
    ("SP-RUN-001", "4012345678905", "Running Shoe AirBoost Pro", "Laufschuhe", "Schwarz/Rot", "44", 89.50, 6, 15),
    ("SP-RUN-002", "4012345679001", "Running Shoe AirBoost Pro", "Laufschuhe", "Weiss/Blau",  "40", 89.50, 6, 15),
    ("SP-RUN-002", "4012345679002", "Running Shoe AirBoost Pro", "Laufschuhe", "Weiss/Blau",  "41", 89.50, 6, 15),
    ("SP-RUN-002", "4012345679003", "Running Shoe AirBoost Pro", "Laufschuhe", "Weiss/Blau",  "42", 89.50, 6, 15),
    ("SP-RUN-002", "4012345679004", "Running Shoe AirBoost Pro", "Laufschuhe", "Weiss/Blau",  "43", 89.50, 6, 15),
    ("SP-FTB-010", "4012345680001", "Fussballschuh Strike Elite FG", "Fussballschuhe", "Gelb/Schwarz", "39", 74.00, 4, 10),
    ("SP-FTB-010", "4012345680002", "Fussballschuh Strike Elite FG", "Fussballschuhe", "Gelb/Schwarz", "40", 74.00, 4, 10),
    ("SP-FTB-010", "4012345680003", "Fussballschuh Strike Elite FG", "Fussballschuhe", "Gelb/Schwarz", "41", 74.00, 4, 10),
    ("SP-FTB-010", "4012345680004", "Fussballschuh Strike Elite FG", "Fussballschuhe", "Gelb/Schwarz", "42", 74.00, 4, 10),
    ("SP-FTB-010", "4012345680005", "Fussballschuh Strike Elite FG", "Fussballschuhe", "Gelb/Schwarz", "43", 74.00, 4, 10),
    ("SP-FTB-010", "4012345680006", "Fussballschuh Strike Elite FG", "Fussballschuhe", "Gelb/Schwarz", "44", 74.00, 4, 10),
    ("SP-JKT-020", "4012345681001", "Trainingsjacke FlexLine", "Bekleidung", "Navy", "S",  42.00, 12, 20),
    ("SP-JKT-020", "4012345681002", "Trainingsjacke FlexLine", "Bekleidung", "Navy", "M",  42.00, 12, 20),
    ("SP-JKT-020", "4012345681003", "Trainingsjacke FlexLine", "Bekleidung", "Navy", "L",  42.00, 12, 20),
    ("SP-JKT-020", "4012345681004", "Trainingsjacke FlexLine", "Bekleidung", "Navy", "XL", 42.00, 12, 20),
    ("SP-JKT-020", "4012345681005", "Trainingsjacke FlexLine", "Bekleidung", "Navy", "XXL",42.00, 12, 20),
    ("SP-JKT-021", "4012345681101", "Trainingsjacke FlexLine", "Bekleidung", "Rot",  "S",  42.00, 12, 20),
    ("SP-JKT-021", "4012345681102", "Trainingsjacke FlexLine", "Bekleidung", "Rot",  "M",  42.00, 12, 20),
    ("SP-JKT-021", "4012345681103", "Trainingsjacke FlexLine", "Bekleidung", "Rot",  "L",  42.00, 12, 20),
    ("SP-JKT-021", "4012345681104", "Trainingsjacke FlexLine", "Bekleidung", "Rot",  "XL", 42.00, 12, 20),
    ("SP-SHT-030", "4012345682001", "Performance T-Shirt DryFit", "Bekleidung", "Weiss", "S",  18.50, 24, 25),
    ("SP-SHT-030", "4012345682002", "Performance T-Shirt DryFit", "Bekleidung", "Weiss", "M",  18.50, 24, 25),
    ("SP-SHT-030", "4012345682003", "Performance T-Shirt DryFit", "Bekleidung", "Weiss", "L",  18.50, 24, 25),
    ("SP-SHT-030", "4012345682004", "Performance T-Shirt DryFit", "Bekleidung", "Weiss", "XL", 18.50, 24, 25),
    ("SP-SHT-030", "4012345682005", "Performance T-Shirt DryFit", "Bekleidung", "Schwarz","S",  18.50, 24, 25),
    ("SP-SHT-030", "4012345682006", "Performance T-Shirt DryFit", "Bekleidung", "Schwarz","M",  18.50, 24, 25),
    ("SP-SHT-030", "4012345682007", "Performance T-Shirt DryFit", "Bekleidung", "Schwarz","L",  18.50, 24, 25),
    ("SP-SHT-030", "4012345682008", "Performance T-Shirt DryFit", "Bekleidung", "Schwarz","XL", 18.50, 24, 25),
    ("SP-BAG-040", "4012345683001", "Sporttasche TeamBag 45L",    "Accessoires","Schwarz","ONE SIZE", 35.00, 6, 10),
    ("SP-BAG-041", "4012345683101", "Sporttasche TeamBag 45L",    "Accessoires","Navy",  "ONE SIZE", 35.00, 6, 10),
    ("SP-BOT-050", "4012345684001", "Trinkflasche HydroSport 750ml","Accessoires","Transparent","ONE SIZE",12.00, 24,  5),
    ("SP-BOT-051", "4012345684101", "Trinkflasche HydroSport 750ml","Accessoires","Blau",   "ONE SIZE",12.00, 24,  5),
]

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
META_FILL     = PatternFill("solid", fgColor="D6E4F0")
ALT_FILL      = PatternFill("solid", fgColor="EBF5FB")
TOTAL_FILL    = PatternFill("solid", fgColor="1F4E79")
TOTAL_FONT    = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

COLUMNS = [
    ("Artikelnummer",  14),
    ("EAN",            18),
    ("Bezeichnung",    30),
    ("Kategorie",      16),
    ("Farbe",          14),
    ("Grösse",         10),
    ("EK-Preis EUR",   14),
    ("Min. Menge",     12),
    ("Rabatt %",       10),
    ("Kollektion",     14),
    ("Lieferdatum",    14),
]


def build():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Offerte"

    # ── Meta rows (rows 1-4) ──────────────────────────────────────────────────
    meta = [
        ("Lieferant:",         "SportPro Distribution AG"),
        ("Währung:",           "EUR"),
        ("Gültig bis:",        "30.06.2025"),
        ("Ansprechpartner:",   "Maria Brunner  |  m.brunner@sportpro.example"),
    ]
    for r, (label, value) in enumerate(meta, start=1):
        ws.cell(r, 1, label).font = Font(bold=True)
        ws.cell(r, 2, value)
        ws.cell(r, 1).fill = META_FILL
        ws.cell(r, 2).fill = META_FILL

    # ── Header row (row 5) ───────────────────────────────────────────────────
    HEADER_ROW = 5
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(HEADER_ROW, col_idx, col_name)
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[HEADER_ROW].height = 28

    # ── Data rows ────────────────────────────────────────────────────────────
    COLLECTION = "Spring/Summer 2025"
    DELIVERY   = "15.03.2025"

    for i, (sku, ean, name, cat, color, size, price, min_qty, disc) in enumerate(PRODUCTS):
        row = HEADER_ROW + 1 + i
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        values = [sku, ean, name, cat, color, size, price, min_qty, disc, COLLECTION, DELIVERY]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row, col_idx, val)
            cell.border = THIN_BORDER
            if fill.fill_type:
                cell.fill = fill
            if col_idx == 7:  # price
                cell.number_format = '#,##0.00 "€"'
            if col_idx == 9:  # discount
                cell.number_format = '0"%"'

    # ── Total / summary row ──────────────────────────────────────────────────
    total_row = HEADER_ROW + 1 + len(PRODUCTS)
    ws.cell(total_row, 1, "Total Positionen").font = TOTAL_FONT
    ws.cell(total_row, 1).fill = TOTAL_FILL
    ws.cell(total_row, 2, len(PRODUCTS)).fill = TOTAL_FILL
    ws.cell(total_row, 2).font = TOTAL_FONT
    for c in range(3, len(COLUMNS) + 1):
        ws.cell(total_row, c).fill = TOTAL_FILL

    # ── Freeze panes + auto-filter ───────────────────────────────────────────
    ws.freeze_panes = f"A{HEADER_ROW + 1}"
    ws.auto_filter.ref = (
        f"A{HEADER_ROW}:{get_column_letter(len(COLUMNS))}{total_row - 1}"
    )

    wb.save(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH}")


if __name__ == "__main__":
    build()
